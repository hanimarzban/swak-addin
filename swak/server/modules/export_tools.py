# cython: language_level=3
"""
SWAK — Data Export Module (11 tools)
Export data to various formats
Files returned as base64 or file path for JS to download
"""

import csv
import json
import io
import base64
import math
import re
from datetime import datetime


def run(tool_id: str, params: dict, headers: list, rows: list) -> dict:
    fn_map = {
        'export-csv':       export_csv,
        'export-json':      export_json,
        'export-xml':       export_xml,
        'export-html':      export_html,
        'export-markdown':  export_markdown,
        'export-sql':       export_sql,
        'export-latex':     export_latex,
        'export-tsv':       export_tsv,
        'export-yaml':      export_yaml,
        'export-excel':     export_excel,
        'export-report':    export_report,
    }
    fn = fn_map.get(tool_id)
    if not fn:
        raise ValueError(f'ابزار ناشناخته: {tool_id}')
    return fn(params, headers, list(rows))


def _result_with_file(title, content_str, filename, mime_type, stats):
    """Return result with base64-encoded file content for JS download"""
    b64 = base64.b64encode(content_str.encode('utf-8')).decode('ascii')
    return {
        'headers': ['filename', 'size', 'format'],
        'rows': [[filename, f'{len(content_str):,} bytes', mime_type]],
        'summary': {'title': title, 'stats': stats},
        'download': {
            'filename': filename,
            'mime_type': mime_type,
            'content_b64': b64,
        }
    }


def _is_empty(v):
    return v is None or (isinstance(v, float) and math.isnan(v)) or (isinstance(v, str) and v.strip() == '')


# ── 1. Export CSV ─────────────────────────────────────────────────────────

def export_csv(params, headers, rows):
    delimiter   = params.get('delimiter', ',')
    encoding    = params.get('encoding', 'utf-8')
    include_bom = params.get('include_bom', 'false') == 'true'
    filename    = params.get('filename', f'export_{_ts()}.csv')

    buf = io.StringIO()
    if include_bom:
        buf.write('\ufeff')  # UTF-8 BOM for Excel compatibility

    writer = csv.writer(buf, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows([['' if _is_empty(v) else v for v in r] for r in rows])
    content = buf.getvalue()

    return _result_with_file('Export CSV', content, filename, 'text/csv',
                             [['ردیف', len(rows)], ['ستون', len(headers)]])


# ── 2. Export JSON ────────────────────────────────────────────────────────

def export_json(params, headers, rows):
    style    = params.get('style', 'records')  # records | array | lines
    indent   = int(params.get('indent', 2))
    filename = params.get('filename', f'export_{_ts()}.json')

    if style == 'records':
        data = [{h: (None if _is_empty(v) else v) for h, v in zip(headers, row)} for row in rows]
        content = json.dumps(data, ensure_ascii=False, indent=indent)
    elif style == 'array':
        content = json.dumps({'headers': headers, 'rows': rows}, ensure_ascii=False, indent=indent)
    else:  # lines (JSONL)
        lines = [json.dumps({h: v for h, v in zip(headers, row)}, ensure_ascii=False) for row in rows]
        content = '\n'.join(lines)

    return _result_with_file('Export JSON', content, filename, 'application/json',
                             [['ردیف', len(rows)], ['فرمت', style]])


# ── 3. Export XML ─────────────────────────────────────────────────────────

def export_xml(params, headers, rows):
    root_tag = params.get('root_tag', 'data')
    row_tag  = params.get('row_tag', 'record')
    filename = params.get('filename', f'export_{_ts()}.xml')

    def escape(v):
        return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    lines = ['<?xml version="1.0" encoding="UTF-8"?>', f'<{root_tag}>']
    for row in rows:
        lines.append(f'  <{row_tag}>')
        for h, v in zip(headers, row):
            tag = re.sub(r'[^a-zA-Z0-9_]', '_', str(h))
            lines.append(f'    <{tag}>{escape(v) if not _is_empty(v) else ""}</{tag}>')
        lines.append(f'  </{row_tag}>')
    lines.append(f'</{root_tag}>')
    content = '\n'.join(lines)

    return _result_with_file('Export XML', content, filename, 'application/xml',
                             [['ردیف', len(rows)], ['تگ ردیف', row_tag]])


# ── 4. Export HTML ────────────────────────────────────────────────────────

def export_html(params, headers, rows):
    title    = params.get('title', 'SWAK Export')
    theme    = params.get('theme', 'dark')  # dark | light
    filename = params.get('filename', f'export_{_ts()}.html')

    bg     = '#0f0f23' if theme == 'dark' else '#ffffff'
    fg     = '#e0e0ff' if theme == 'dark' else '#111111'
    hdr_bg = '#1a1a3e' if theme == 'dark' else '#2563eb'
    hdr_fg = '#a0a0ff' if theme == 'dark' else '#ffffff'
    alt_bg = '#0a0a1a' if theme == 'dark' else '#f1f5f9'
    border = '#2a2a5a' if theme == 'dark' else '#e2e8f0'

    def escape(v):
        return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    rows_html = ''
    for i, row in enumerate(rows):
        bg_row = alt_bg if i % 2 == 0 else bg
        cells  = ''.join(f'<td style="padding:6px 12px;border-bottom:1px solid {border}">{escape(v) if not _is_empty(v) else ""}</td>'
                         for v in row)
        rows_html += f'<tr style="background:{bg_row}">{cells}</tr>\n'

    headers_html = ''.join(f'<th style="padding:8px 12px;text-align:left;background:{hdr_bg};color:{hdr_fg};font-weight:600">{escape(h)}</th>'
                           for h in headers)

    content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{escape(title)}</title>
  <style>
    * {{ box-sizing:border-box; margin:0; padding:0 }}
    body {{ font-family:Tahoma,Arial,sans-serif; background:{bg}; color:{fg}; padding:20px }}
    h1 {{ margin-bottom:16px; font-size:1.4rem; color:{hdr_fg if theme=='dark' else '#1e40af'} }}
    .meta {{ font-size:0.8rem; color:#666; margin-bottom:12px }}
    table {{ border-collapse:collapse; width:100%; font-size:0.88rem }}
    thead tr {{ background:{hdr_bg} }}
    tr:hover {{ opacity:0.9 }}
  </style>
</head>
<body>
  <h1>{escape(title)}</h1>
  <div class="meta">Generated by SWAK v2.0 • {len(rows):,} rows × {len(headers)} columns • {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
  <table>
    <thead><tr>{headers_html}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</body>
</html>"""

    return _result_with_file('Export HTML', content, filename, 'text/html',
                             [['ردیف', len(rows)], ['theme', theme]])


# ── 5. Export Markdown ────────────────────────────────────────────────────

def export_markdown(params, headers, rows):
    filename  = params.get('filename', f'export_{_ts()}.md')
    title     = params.get('title', '')
    max_width = int(params.get('max_col_width', 30))

    def trunc(v, n):
        s = str(v) if not _is_empty(v) else ''
        return s[:n] + '…' if len(s) > n else s

    col_widths = [max(len(h), max((len(trunc(r[i] if i < len(r) else '', max_width))
                  for r in rows), default=0)) for i, h in enumerate(headers)]
    col_widths = [max(w, 3) for w in col_widths]

    def fmt_row(vals):
        cells = [str(trunc(vals[i] if i < len(vals) else '', max_width)).ljust(col_widths[i])
                 for i in range(len(headers))]
        return '| ' + ' | '.join(cells) + ' |'

    separator = '| ' + ' | '.join('-' * w for w in col_widths) + ' |'
    lines     = []
    if title:
        lines.extend([f'# {title}', '', f'*{len(rows)} rows × {len(headers)} columns*', ''])
    lines.append(fmt_row(headers))
    lines.append(separator)
    lines.extend(fmt_row(row) for row in rows)
    content = '\n'.join(lines)

    return _result_with_file('Export Markdown', content, filename, 'text/markdown',
                             [['ردیف', len(rows)], ['ستون', len(headers)]])


# ── 6. Export SQL INSERT ──────────────────────────────────────────────────

def export_sql(params, headers, rows):
    table_name = params.get('table_name', 'data')
    db_type    = params.get('db_type', 'postgresql')
    filename   = params.get('filename', f'export_{_ts()}.sql')
    batch_size = int(params.get('batch_size', 100))

    def quote_val(v):
        if _is_empty(v): return 'NULL'
        if isinstance(v, (int, float)): return str(v)
        s = str(v).replace("'", "''")
        return f"'{s}'"

    def quote_col(h):
        return f'"{h}"' if db_type in ('postgresql','sqlite') else f'`{h}`'

    cols  = ', '.join(quote_col(h) for h in headers)
    lines = [f'-- Generated by SWAK v2.0 — {datetime.now().strftime("%Y-%m-%d %H:%M")}',
             f'-- Table: {table_name} ({len(rows)} rows)',
             '']

    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        vals  = ',\n  '.join('(' + ', '.join(quote_val(v) for v in row) + ')' for row in batch)
        lines.append(f'INSERT INTO {quote_col(table_name)} ({cols})')
        lines.append(f'VALUES')
        lines.append(f'  {vals};')
        lines.append('')

    content = '\n'.join(lines)
    return _result_with_file('Export SQL', content, filename, 'application/sql',
                             [['ردیف', len(rows)], ['db_type', db_type], ['batch_size', batch_size]])


# ── 7. Export LaTeX Table ─────────────────────────────────────────────────

def export_latex(params, headers, rows):
    filename = params.get('filename', f'export_{_ts()}.tex')
    caption  = params.get('caption', 'Data Table')
    label    = params.get('label', 'tab:data')

    def escape_latex(v):
        s = str(v) if not _is_empty(v) else ''
        replacements = [
            ('\\', '\\textbackslash{}'),
            ('&',  '\\&'),
            ('%',  '\\%'),
            ('$',  '\\$'),
            ('#',  '\\#'),
            ('^',  '\\textasciicircum{}'),
            ('~',  '\\textasciitilde{}'),
            ('{',  '\\{'),
            ('}',  '\\}'),
            ('_',  '\\_'),
        ]
        for ch, rep in replacements:
            s = s.replace(ch, rep)
        return s

    col_fmt  = 'l' * len(headers)
    hdr_row  = ' & '.join(f'\\textbf{{{escape_latex(h)}}}' for h in headers) + ' \\\\'
    sep      = '\\hline'
    data_rows= ' \\\\\n'.join(' & '.join(escape_latex(v) for v in row) for row in rows)

    content = f"""\\begin{{table}}[h]
\\centering
\\caption{{{escape_latex(caption)}}}
\\label{{{label}}}
\\begin{{tabular}}{{{col_fmt}}}
\\hline
{hdr_row}
{sep}
{data_rows} \\\\
\\hline
\\end{{tabular}}
\\end{{table}}"""

    return _result_with_file('Export LaTeX', content, filename, 'text/plain',
                             [['ردیف', len(rows)], ['caption', caption]])


# ── 8. Export TSV ─────────────────────────────────────────────────────────

def export_tsv(params, headers, rows):
    filename = params.get('filename', f'export_{_ts()}.tsv')
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter='\t')
    writer.writerow(headers)
    writer.writerows([['' if _is_empty(v) else v for v in r] for r in rows])
    content = buf.getvalue()

    return _result_with_file('Export TSV', content, filename, 'text/tab-separated-values',
                             [['ردیف', len(rows)]])


# ── 9. Export YAML ────────────────────────────────────────────────────────

def export_yaml(params, headers, rows):
    filename = params.get('filename', f'export_{_ts()}.yaml')

    def yaml_val(v):
        if _is_empty(v): return 'null'
        if isinstance(v, bool): return 'true' if v else 'false'
        if isinstance(v, (int, float)): return str(v)
        s = str(v)
        if any(c in s for c in [':', '#', '[', ']', '{', '}', ',', '&', '*', '?', '|', '-', '<', '>', '!', "'", '"', '%', '@', '`', '\n']):
            return f'"{s.replace(chr(34), chr(92)+chr(34))}"'
        return s

    lines = ['# Generated by SWAK v2.0', f'# {len(rows)} records', 'data:']
    for row in rows:
        lines.append('  - ' + '\n    '.join(
            ('' if i == 0 else '') + f'{re.sub(r"[^a-zA-Z0-9_]","_",h)}: {yaml_val(v)}'
            for i, (h, v) in enumerate(zip(headers, row))
        ))
    content = '\n'.join(lines)

    return _result_with_file('Export YAML', content, filename, 'application/yaml',
                             [['ردیف', len(rows)]])


# ── 10. Export Excel (xlsx via openpyxl if available) ────────────────────

def export_excel(params, headers, rows):
    filename   = params.get('filename', f'export_{_ts()}.xlsx')
    sheet_name = params.get('sheet_name', 'Sheet1')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header style
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='1A1A3E')

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for ri, row in enumerate(rows, 2):
            for ci, v in enumerate(row, 1):
                ws.cell(ri, ci, None if _is_empty(v) else v)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode('ascii')

        return {
            'headers': ['filename', 'size', 'format'],
            'rows': [[filename, f'{buf.tell():,} bytes', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']],
            'summary': {'title': 'Export Excel', 'stats': [['ردیف', len(rows)], ['ستون', len(headers)]]},
            'download': {'filename': filename,
                         'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         'content_b64': b64}
        }

    except ImportError:
        # Fallback: export as CSV with .xlsx extension note
        return export_csv({'filename': filename.replace('.xlsx', '.csv')}, headers, rows)


# ── 11. Export Full Report (HTML) ─────────────────────────────────────────

def export_report(params, headers, rows):
    title     = params.get('title', 'SWAK Data Report')
    filename  = params.get('filename', f'report_{_ts()}.html')
    include_stats = params.get('include_stats', 'true') == 'true'

    # Compute summary stats for numeric columns
    def col_stats(ci):
        vals = [v for r in rows for v in [r[ci] if ci < len(r) else None]
                if isinstance(v, (int, float)) and not math.isnan(v)]
        if not vals: return None
        s = sorted(vals)
        m = sum(vals)/len(vals)
        return {'count': len(vals), 'min': min(vals), 'max': max(vals),
                'mean': round(m, 2), 'median': s[len(s)//2]}

    stats_html = ''
    if include_stats:
        stats_rows = ''
        for i, h in enumerate(headers):
            st = col_stats(i)
            if st:
                stats_rows += f'<tr><td>{h}</td><td>{st["count"]}</td><td>{st["min"]}</td><td>{st["max"]}</td><td>{st["mean"]}</td><td>{st["median"]}</td></tr>'
        if stats_rows:
            stats_html = f'''
<h2 style="margin:20px 0 10px;color:#a0a0ff">آمار توصیفی</h2>
<table>
<thead><tr><th>ستون</th><th>تعداد</th><th>کمینه</th><th>بیشینه</th><th>میانگین</th><th>میانه</th></tr></thead>
<tbody>{stats_rows}</tbody>
</table>'''

    # Reuse HTML export for data table
    html_result = export_html({'title': '', 'theme': 'dark'}, headers, rows[:1000])
    # Extract table HTML
    table_match = re.search(r'<table.*</table>', html_result['download']['content_b64'], re.DOTALL)

    from datetime import datetime
    content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<style>
* {{box-sizing:border-box;margin:0;padding:0}}
body {{font-family:Tahoma,Arial,sans-serif;background:#0f0f23;color:#e0e0ff;padding:24px;direction:rtl}}
h1 {{font-size:1.6rem;color:#a0a0ff;margin-bottom:8px}}
h2 {{font-size:1.1rem;color:#8080dd}}
.meta {{color:#666;font-size:0.8rem;margin-bottom:20px}}
table {{border-collapse:collapse;width:100%;font-size:0.85rem;margin-bottom:24px}}
th {{background:#1a1a3e;color:#a0a0ff;padding:8px 12px;text-align:right;font-weight:600}}
td {{padding:6px 12px;border-bottom:1px solid #2a2a5a}}
tr:nth-child(even) {{background:#0a0a1a}}
</style>
</head>
<body>
<h1>📊 {title}</h1>
<div class="meta">تولید شده توسط SWAK v2.0 • {len(rows):,} ردیف × {len(headers)} ستون • {datetime.now().strftime("%Y-%m-%d %H:%M")}</div>
{stats_html}
<h2 style="margin:20px 0 10px;color:#a0a0ff">داده‌ها{' (اولین ۱۰۰۰ ردیف)' if len(rows) > 1000 else ''}</h2>
<table>
<thead><tr>{''.join(f"<th>{h}</th>" for h in headers)}</tr></thead>
<tbody>{''.join('<tr>' + ''.join(f"<td>{'' if _is_empty(v) else v}</td>" for v in row) + '</tr>' for row in rows[:1000])}</tbody>
</table>
</body>
</html>"""

    return _result_with_file('Export Report', content, filename, 'text/html',
                             [['ردیف', len(rows)], ['آمار', 'بله' if include_stats else 'خیر']])


def _ts():
    return datetime.now().strftime('%Y%m%d_%H%M%S')
